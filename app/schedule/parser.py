from __future__ import annotations

import re
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.schedule.models import Lesson, Schedule

DATE_RE = re.compile(r"(?P<d>\d{1,2})\.(?P<m>\d{1,2})\.(?P<y>\d{4})")
PAIR_RE = re.compile(
    r"(?P<pair>\d+)\s+.*?(?P<sh>\d{1,2}):(?P<sm>\d{2})\s*[-–—]\s*(?P<eh>\d{1,2}):(?P<em>\d{2})",
    re.S,
)
URL_RE = re.compile(r"https?://[^\s]+", re.I)
LOCATION_RE = re.compile(r"\(([^()]*(?:онлайн|\d{2,4}\s*\[\d+\])[^()]*)\)\s*$", re.I)
TEACHER_RE = re.compile(r"([А-ЯЁ][а-яё-]+(?:[ \t]+[А-ЯЁ][а-яё-]+)*[ \t]+[А-ЯЁ]\.[А-ЯЁ]\.)")
NOTE_RE = re.compile(r"\(([А-ЯЁA-Z]{2,8})\)")
LESSON_TYPES: tuple[tuple[re.Pattern[str], str], ...] = (
    (
        re.compile(
            r"\((?:лек(?:ц(?:ия|ии)?)?\.?|л)\)|\bлек(?:ц(?:ия|ии)?)?\.|\bлекц(?:ия|ии)\b",
            re.I,
        ),
        "лекция",
    ),
    (re.compile(r"\((?:семинар|сем\.?|с)\)|\bсеминар\w*\b", re.I), "семинар"),
    (re.compile(r"\bпракт(?:ика|ическое|\.)\s*(?:занятие|курс)?", re.I), "практика"),
    (re.compile(r"\bлаб(?:ораторная|\.)\s*(?:работа)?", re.I), "лабораторная"),
    (re.compile(r"\bэкзамен\w*\b", re.I), "экзамен"),
    (re.compile(r"\bзач[её]т\w*\b", re.I), "зачёт"),
    (re.compile(r"\bконсультац\w*\b", re.I), "консультация"),
)
GROUP_RE = re.compile(r"^[А-ЯЁA-Z]{1,8}-\d{2}-\d+$")
COURSE_RE = re.compile(r"^\s*([1-9])\s+курс\s*$", re.I)


class ScheduleParseError(Exception):
    """Workbook cannot be interpreted as a schedule."""


def merged_value(sheet: Worksheet, row: int, column: int) -> Any:
    cell = sheet.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col).value
    return None


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    match = DATE_RE.search(str(value or ""))
    return date(int(match["y"]), int(match["m"]), int(match["d"])) if match else None


def parse_pair(value: object) -> tuple[int, time, time] | None:
    match = PAIR_RE.search(str(value or "").strip())
    if not match:
        return None
    return (
        int(match["pair"]),
        time(int(match["sh"]), int(match["sm"])),
        time(int(match["eh"]), int(match["em"])),
    )


def parse_lesson_text(
    text: str,
) -> tuple[str, str | None, str | None, bool, str | None, tuple[str, ...], str | None]:
    value = text.strip()
    url_match = URL_RE.search(value)
    url = url_match.group(0).rstrip(".,)") if url_match else None
    lines = [
        line.strip() for line in value.splitlines() if line.strip() and not line.startswith("http")
    ]
    teacher_match = TEACHER_RE.search(value)
    teacher = teacher_match.group(1).strip() if teacher_match else None
    location_match = LOCATION_RE.search(value)
    location_raw = location_match.group(1).strip() if location_match else None
    is_online = bool(re.search(r"онлайн", value, re.I))
    location = "онлайн" if is_online else location_raw
    notes = tuple(dict.fromkeys(NOTE_RE.findall(value)))
    found_types: list[tuple[int, str]] = []
    for pattern, normalized in LESSON_TYPES:
        found_types.extend((match.start(), normalized) for match in pattern.finditer(value))
    lesson_types = dict.fromkeys(
        normalized for _, normalized in sorted(found_types, key=lambda item: item[0])
    )
    lesson_type = " / ".join(lesson_types) or None
    subject_parts: list[str] = []
    for line in lines:
        if teacher and teacher in line:
            prefix = line[: line.index(teacher)].strip()
            prefix = NOTE_RE.sub("", prefix).strip()
            if prefix:
                subject_parts.append(prefix)
            break
        if location_match and line == location_match.group(0):
            break
        subject_parts.append(line)
    subject = " ".join(subject_parts).strip()
    subject = NOTE_RE.sub("", subject).strip()
    if lesson_type:
        subject = re.sub(
            r"\s*\((?:лекц(?:ия|ии)?|л|семинар|сем\.?|с)\)\s*$", "", subject, flags=re.I
        )
        subject = re.sub(
            r"\s+(?:семинар|лекция|экзамен|зач[её]т|консультация)\s*$",
            "",
            subject,
            flags=re.I,
        ).strip(" .—-")
    subject = re.sub(r"\bинформац\.\s+систем\b", "информационных систем", subject, flags=re.I)
    if not subject:
        subject = lines[0] if lines else value
    return subject, teacher, location, is_online, url, notes, lesson_type


class ExcelScheduleParser:
    def parse(self, source: str | Path | bytes | BinaryIO) -> Schedule:
        stream: str | Path | BinaryIO
        stream = BytesIO(source) if isinstance(source, bytes) else source
        workbook = load_workbook(stream, data_only=True)
        courses: dict[int, tuple[str, ...]] = {}
        lessons: list[Lesson] = []
        for sheet in workbook.worksheets:
            course_match = COURSE_RE.match(sheet.title)
            if not course_match:
                continue
            course = int(course_match.group(1))
            header_row, group_columns = self._find_groups(sheet)
            courses[course] = tuple(dict.fromkeys(group_columns.values()))
            current_date: date | None = None
            for row in range(header_row + 1, sheet.max_row + 1):
                current_date = parse_date(merged_value(sheet, row, 1)) or current_date
                pair = parse_pair(merged_value(sheet, row, 2))
                if current_date is None or pair is None:
                    continue
                for column, group in group_columns.items():
                    raw = merged_value(sheet, row, column)
                    if not isinstance(raw, str) or not raw.strip():
                        continue
                    subject, teacher, location, online, url, notes, lesson_type = parse_lesson_text(
                        raw
                    )
                    lessons.append(
                        Lesson(
                            group,
                            current_date,
                            *pair,
                            subject,
                            teacher,
                            location,
                            online,
                            url,
                            notes,
                            lesson_type,
                        )
                    )
        if not courses:
            raise ScheduleParseError("No course sheets with group headers found")
        unique = tuple(dict.fromkeys(lessons))
        return Schedule(courses, unique)

    @staticmethod
    def _find_groups(sheet: Worksheet) -> tuple[int, dict[int, str]]:
        for row in range(1, min(sheet.max_row, 15) + 1):
            found: dict[int, str] = {}
            for column in range(3, sheet.max_column + 1):
                value = merged_value(sheet, row, column)
                text = str(value or "").strip()
                if GROUP_RE.match(text):
                    found[column] = text
            if found:
                return row, found
        raise ScheduleParseError(f"No group header found in sheet {sheet.title!r}")
