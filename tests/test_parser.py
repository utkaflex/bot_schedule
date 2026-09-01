from datetime import date, time
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.schedule.parser import (
    ExcelScheduleParser,
    ScheduleParseError,
    merged_value,
    parse_date,
    parse_lesson_text,
    parse_pair,
)


def workbook_bytes(*, merged: bool = False) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "4 курс"
    sheet.append(["День", "Пара", "РИС-23-1", "РИС-23-2"])
    sheet.append(
        ["Вторник\n01.09.2026", "3\n\n11:30-12:50", "Предмет\n\nИванов И.И. (115[2])", None]
    )
    if merged:
        sheet.merge_cells("C2:D2")
    output = BytesIO()
    book.save(output)
    return output.getvalue()


@pytest.mark.parametrize(
    ("raw", "expected"), [("Вторник\n01.09.2026", date(2026, 9, 1)), ("", None), (None, None)]
)
def test_parse_date(raw, expected):
    assert parse_date(raw) == expected


def test_parse_pair():
    assert parse_pair("3\n\n11:30–12:50") == (3, time(11, 30), time(12, 50))


def test_invalid_pair():
    assert parse_pair("перерыв") is None


def test_offline_lesson_fields():
    result = parse_lesson_text("Автоматизация\n\nСуворов А.О. (115[2])")
    assert result == ("Автоматизация", "Суворов А.О.", "115[2]", False, None, (), None)


def test_online_url_and_notes():
    result = parse_lesson_text(
        "Управление проектами\n\n(МКД) Климов Б.А. (онлайн[0])\nhttps://meet.test/a"
    )
    assert result == (
        "Управление проектами",
        "Климов Б.А.",
        "онлайн",
        True,
        "https://meet.test/a",
        ("МКД",),
        None,
    )


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("Анализ данных (лекции)\n\nИванов И.И. (101[1])", "лекция"),
        ("Анализ данных (Л)\n\nИванов И.И. (101[1])", "лекция"),
        ("Анализ данных семинар\n\nИванов И.И. (101[1])", "семинар"),
        ("Практ. курс английского языка\n\nИванов И.И. (101[1])", "практика"),
        ("Экзамен по анализу данных\n\nИванов И.И. (101[1])", "экзамен"),
    ],
)
def test_lesson_type(text, expected):
    assert parse_lesson_text(text)[-1] == expected


def test_detects_groups_and_normal_cell():
    schedule = ExcelScheduleParser().parse(workbook_bytes())
    assert schedule.courses == {4: ("РИС-23-1", "РИС-23-2")}
    assert len(schedule.for_group("РИС-23-1")) == 1
    assert schedule.for_group("РИС-23-2") == ()


def test_merged_cell_applies_to_every_group():
    schedule = ExcelScheduleParser().parse(workbook_bytes(merged=True))
    assert [x.group for x in schedule.lessons] == ["РИС-23-1", "РИС-23-2"]


def test_merged_value_uses_top_left():
    book = Workbook()
    sheet = book.active
    sheet["C6"] = "shared"
    sheet.merge_cells("C6:F6")
    assert merged_value(sheet, 6, 5) == "shared"


def test_rejects_non_schedule_workbook():
    book = Workbook()
    output = BytesIO()
    book.save(output)
    with pytest.raises(ScheduleParseError):
        ExcelScheduleParser().parse(output.getvalue())


def test_real_schedule_all_ten_lessons():
    lessons = ExcelScheduleParser().parse("tests/fixtures/real_schedule.xlsx").for_group("РИС-23-3")
    expected = [
        (1, "11:30", "Автоматизация деятельности предприятия", "Суворов А.О.", "115[2]", False, ()),
        (1, "13:10", "Автоматизация деятельности предприятия", "Суворов А.О.", "115[2]", False, ()),
        (
            2,
            "09:40",
            "Модельно-ориентированная разработка информационных систем",
            "Лядова Л.Н.",
            "401[1]",
            False,
            (),
        ),
        (2, "11:30", "Распознавание образов", "Замятина Е.Б.", "401[1]", False, ()),
        (
            2,
            "13:10",
            "Средства BI для управления процессами и задачами",
            "Комиссаров К.А.",
            "401[1]",
            False,
            (),
        ),
        (3, "13:10", "Предпринимательство", "Францкевич А.Р.", "403[1]", False, ()),
        (3, "15:00", "Предпринимательство", "Францкевич А.Р.", "403[1]", False, ()),
        (4, "09:40", "Разработка AI-агентов", "Ланин В.В.", "онлайн", True, ()),
        (4, "11:30", "Разработка AI-агентов", "Ланин В.В.", "онлайн", True, ()),
        (5, "11:30", "Управление программными проектами", "Климов Б.А.", "онлайн", True, ("МКД",)),
    ]
    actual = [
        (
            x.date.day,
            x.start_time.strftime("%H:%M"),
            x.subject,
            x.teacher,
            x.location,
            x.is_online,
            x.notes,
        )
        for x in lessons
    ]
    assert actual == expected
